from dataclasses import dataclass
from typing import Callable

import numpy as np
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet

from graph_generator.dto.ToGraph import ToGraph
from graph_generator.dto.toml import Filter
from graph_generator.parser import ParsedTable


@dataclass
class Bounds:
    start: int
    end: int


@dataclass
class ColumnLabels:
    x_bounds: Bounds
    y: dict[str, int]
    sorted_keys: list[str]


@dataclass
class RowLabels:
    y_bounds: Bounds
    x: dict[str, int]
    sorted_keys: list[str]


@dataclass
class TableInfo:
    column_labels: ColumnLabels
    row_labels: RowLabels


def get_index_in_column(sheet: Worksheet, column: int, first_row: int, last_row: int, value: str) -> int | None:
    """Returns the index of the first occurrence of value in the specified column of the sheet.
    If the value is not found, returns None.
    """
    for row in range(first_row, last_row + 1):
        cell_value = str(sheet.cell(row=row, column=column).value)
        if cell_value == value:
            return row

    return None


def get_index_in_row(sheet: Worksheet, row: int, first_col: int, last_col: int, value: str) -> int | None:
    """Returns the index of the first occurrence of value in the specified row of the sheet.
    If the value is not found, returns None.
    """
    for col in range(first_col, last_col + 1):
        cell_value = str(sheet.cell(row=row, column=col).value)
        if cell_value == value:
            return col

    return None


def craft_table_info(sheet: Worksheet, table: ParsedTable) -> TableInfo:
    # order labels
    ordered_column_labels = sorted(table.column_header.keys(), key=lambda item: table.column_header[item])
    ordered_row_labels = sorted(table.row_header.keys(), key=lambda item: table.row_header[item])

    # find top row of data
    last_column_header = table.column_header[ordered_column_labels[-1]]
    top = last_column_header + 1

    # find left column of data
    last_row_header = table.row_header[ordered_row_labels[-1]]
    left = last_row_header + 1

    # check for last row
    last_row = top
    while sheet.cell(row=last_row + 1, column=left).value is not None:
        last_row += 1

    # check for last column
    last_column = left
    while sheet.cell(row=top, column=last_column + 1).value is not None:
        last_column += 1

    return TableInfo(
        column_labels=ColumnLabels(Bounds(left, last_column), {key: col for key, col in table.column_header.items()}, ordered_column_labels),
        row_labels=RowLabels(Bounds(top, last_row), {key: row for key, row in table.row_header.items()}, ordered_row_labels),
    )


def load_all_row_labels(sheet: Worksheet, table_info: TableInfo, column: str) -> list[str]:
    row_labels = []
    for row in range(table_info.row_labels.y_bounds.start, table_info.row_labels.y_bounds.end + 1):
        cell_value = sheet.cell(row=row, column=table_info.row_labels.x[column]).value
        if cell_value is not None:
            row_labels.append(str(cell_value))

    pattern_length = find_pattern_length(row_labels)
    if pattern_length is not None:
        return row_labels[:pattern_length]

    return row_labels


def find_pattern_length(lst: list[str]) -> int | None:
    n = len(lst)

    if n < 2:
        return None

    first = lst[0]

    # find next occurrence of first element
    for i in range(1, n):
        if lst[i] == first:
            return i

    return None


def load_all_column_labels(sheet: Worksheet, table_info: TableInfo, row: str) -> list[str]:
    column_labels = []
    for col in range(table_info.column_labels.x_bounds.start, table_info.column_labels.x_bounds.end + 1):
        cell_value = sheet.cell(row=table_info.column_labels.y[row], column=col).value
        if cell_value is not None:
            column_labels.append(str(cell_value))

    pattern_length = find_pattern_length(column_labels)
    if pattern_length is not None:
        return column_labels[:pattern_length]

    return column_labels


def complete_info(what: dict[str, list[str]] | None, sorted_keys: list[str], load: Callable[[str], list[str]]) -> dict[str, list[str]]:
    labels = {}

    if what is None:
        for key in sorted_keys:
            labels[key] = load(key)
    else:
        for key in sorted_keys:
            if key in what:
                labels[key] = what[key]
            else:
                labels[key] = load(key)

    return labels


def load_data(table: ParsedTable, filter: Filter) -> ToGraph:
    book = openpyxl.load_workbook(filename=table.source_file)
    sheet_num = book.sheetnames.index(table.sheet)

    if sheet_num == -1:
        raise ValueError(f"Sheet {table.sheet} not found in {table.source_file}")

    sheet = book.worksheets[sheet_num]

    table_info = craft_table_info(sheet, table)

    row_labels = complete_info(filter.row, table_info.row_labels.sorted_keys, lambda key: load_all_row_labels(sheet, table_info, key))
    column_labels = complete_info(filter.column, table_info.column_labels.sorted_keys, lambda key: load_all_column_labels(sheet, table_info, key))

    data = traverse_row_labels(sheet, table_info, row_labels, column_labels, table_info.row_labels.y_bounds.start)

    # TODO: warn when there are non-numeric values
    return ToGraph(
        [row_labels[key] for key in table_info.row_labels.sorted_keys] + [column_labels[key] for key in table_info.column_labels.sorted_keys], np.array(data)
    )


def traverse(sheet: Worksheet, row: int, table_info: TableInfo, series: dict[str, list[str]], x: int, depth=0):
    if depth == len(series):
        return sheet.cell(row=row, column=x).value

    result = []
    key = table_info.column_labels.sorted_keys[depth]
    for item in series[key]:
        new_x = get_index_in_row(sheet, table_info.column_labels.y[key], x, table_info.column_labels.x_bounds.end, item)

        if new_x is None:
            raise ValueError(f"Column label {item} not found in table")

        result.append(traverse(sheet, row, table_info, series, new_x, depth + 1))

    return result


def traverse_row_labels(sheet: Worksheet, table_info: TableInfo, series: dict[str, list[str]], column_series: dict[str, list[str]], y: int, depth=0):
    if depth == len(series):
        return traverse(sheet, y, table_info, column_series, table_info.column_labels.x_bounds.start)

    result = []
    key = table_info.row_labels.sorted_keys[depth]
    for item in series[key]:
        new_y = get_index_in_column(sheet, table_info.row_labels.x[key], y, table_info.row_labels.y_bounds.end, item)

        if new_y is None:
            raise ValueError(f"Row label {item} not found in table at column {key}")

        result.append(traverse_row_labels(sheet, table_info, series, column_series, new_y, depth + 1))

    return result
